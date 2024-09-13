import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Ocultar a janela principal do Tkinter
Tk().withdraw()

# Abrir o explorador de arquivos para selecionar o CSV
file_path = askopenfilename(
    title="Selecione o arquivo CSV",
    filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
)

if not file_path:
    print("Nenhum arquivo foi selecionado.")
else:
    try:
        # Ler o arquivo CSV completo
        df = pd.read_csv(
            file_path,
            usecols=['Nome_Completo', 'Paginas_Color', 'Paginas_Mono', 'Data_de_Impressão'],
            encoding='ISO-8859-1',
            delimiter=';',
            on_bad_lines='skip'
        )
        
        # Remover espaços em branco e caracteres estranhos
        df['Data_de_Impressão'] = df['Data_de_Impressão'].str.strip()
        
        # Converter a coluna 'Data_de_Impressão' para o tipo datetime
        df['Data_de_Impressão'] = pd.to_datetime(df['Data_de_Impressão'], format='%d/%m/%Y %H:%M', errors='coerce')
        
        # Definir o intervalo de datas
        start_date = pd.to_datetime('07/08/2024', format='%d/%m/%Y')
        end_date = pd.to_datetime('07/09/2024', format='%d/%m/%Y')
        
        # Filtrar os dados entre as duas datas
        df_filtered = df[(df['Data_de_Impressão'] >= start_date) & (df['Data_de_Impressão'] <= end_date)]
        
        # Criar a coluna 'pagina' como a soma de 'Paginas_Mono' e 'Paginas_Color'
        df_filtered['pagina'] = df_filtered['Paginas_Color'] + df_filtered['Paginas_Mono']
        
        # Renomear as colunas
        df_filtered = df_filtered.rename(columns={
            'Nome_Completo': 'Nome',
            'Paginas_Color': 'Colorido',
            'Paginas_Mono': 'P&B',
            'pagina': 'Pagina'
        })
        
        # Agrupar pela coluna 'Nome' e somar as colunas numéricas
        df_agrupado = df_filtered.groupby('Nome').agg({
            'Colorido': 'sum',
            'P&B': 'sum',
            'Pagina': 'sum'
        }).reset_index()
        
        # Adicionar a linha de total
        total_row = pd.DataFrame(df_agrupado[['Colorido', 'P&B', 'Pagina']].sum()).T
        total_row['Nome'] = 'Total'
        df_agrupado = pd.concat([df_agrupado, total_row], ignore_index=True)
        
        # Salvando o resultado em um arquivo Excel
        output_file = 'resultado.xlsx'
        df_agrupado.to_excel(output_file, index=False)
        
        # Aplicar bordas e formatação às células no Excel
        wb = load_workbook(output_file)
        ws = wb.active
        
        # Definir bordas
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        
        # Definir fonte em negrito
        bold_font = Font(bold=True)
        
        # Aplicar bordas e fonte em negrito ao cabeçalho
        for cell in ws[1]:
            cell.border = border
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Aplicar bordas e fonte em negrito à linha de total
        total_row_index = len(df_agrupado) + 1
        for cell in ws[total_row_index]:
            cell.border = border
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Aplicar bordas a todas as outras células da planilha
        for row in ws.iter_rows(min_row=2, max_row=total_row_index-1, min_col=1, max_col=len(df_agrupado.columns)):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Salvar o arquivo com bordas e formatação
        wb.save(output_file)

        print("Operação concluída! Dados filtrados, unificados, totalizados e salvos em Excel com bordas e formatação.")
    
    except pd.errors.ParserError as e:
        print(f'Erro ao ler o CSV: {e}')
    except Exception as e:
        print(f'Erro ao processar o arquivo: {e}')
